#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""将Excel文件拆分成多列，每列100条记录，每列从第1行开始"""

import openpyxl
import os
import math

def split_excel_to_columns(input_file, rows_per_column=100):
    """将Excel文件的单列数据拆分成多列，每列从第1行开始"""
    
    # 读取输入文件
    wb_input = openpyxl.load_workbook(input_file)
    ws_input = wb_input.active
    
    # 获取所有数据
    all_data = []
    for row in ws_input.iter_rows(values_only=True):
        if row[0] is not None:
            all_data.append(row[0])
    
    total_rows = len(all_data)
    print(f"总共 {total_rows} 条数据")
    
    # 计算需要的列数
    num_columns = math.ceil(total_rows / rows_per_column)
    print(f"将拆分成 {num_columns} 列（每列 {rows_per_column} 条）")
    
    # 创建新工作簿
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    ws_output.title = "拆分结果"
    
    # 填充数据：每列从第1行开始
    for col_idx in range(num_columns):
        start_idx = col_idx * rows_per_column
        end_idx = min(start_idx + rows_per_column, total_rows)
        
        # 在新列中填充数据，从第1行开始
        for data_idx in range(start_idx, end_idx):
            ws_output.cell(row=data_idx - start_idx + 1, column=col_idx + 1, value=all_data[data_idx])
        
        print(f"列 {col_idx + 1}: 第1行 ~ 第{end_idx - start_idx}行（共{end_idx - start_idx}条）")
    
    # 保存文件
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    dir_name = os.path.dirname(input_file)
    output_file = os.path.join(dir_name, f"{base_name}_拆分_v2.xlsx")
    
    wb_output.save(output_file)
    wb_input.close()
    
    print(f"\n✅ 完成！文件已保存到: {output_file}")
    return output_file

if __name__ == "__main__":
    input_file = r"D:\Users\Administrator\Desktop\下架.xlsx"
    split_excel_to_columns(input_file, rows_per_column=100)
