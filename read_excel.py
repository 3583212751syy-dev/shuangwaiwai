#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""查看Excel文件内容的脚本"""

import openpyxl
import sys

def read_excel_info(file_path):
    """读取Excel文件信息"""
    try:
        wb = openpyxl.load_workbook(file_path)
        print(f"工作表名称: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n=== 工作表: {sheet_name} ===")
            print(f"行数: {ws.max_row}")
            print(f"列数: {ws.max_column}")
            
            # 显示前5行数据
            print("\n前5行数据:")
            for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
                print(f"第{row_idx}行: {row[:10]}")  # 只显示前10列
            
            # 显示列标题（如果有的话）
            print("\n列标题（第1行）:")
            headers = []
            for col in range(1, min(ws.max_column + 1, 21)):  # 最多显示20列
                cell_value = ws.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value else f"列{col}")
            print(headers)
            
        wb.close()
        
    except Exception as e:
        print(f"错误: {e}")

if __name__ == "__main__":
    file_path = r"D:\Users\Administrator\Desktop\下架.xlsx"
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    read_excel_info(file_path)
