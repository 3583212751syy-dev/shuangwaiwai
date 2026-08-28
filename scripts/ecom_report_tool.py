# -*- coding: utf-8 -*-
"""
ecom_report_tool.py — 电商产品图片分析报表工具 v3.0
=====================================================
v3 变化：
1. 删除空运方案（用户：报表只算海运；mixed=首轮海运+海运补货）
2. 物流按克重计费：海运 ¥3/kg × 克重；印尼段 ≤500g ¥2/件、>500g ¥4/kg（区域费率库）
3. 币种可自主选择（--currency），金额双币种显示（当地币 + 自动换算人民币）
4. 区域物流费率：按平台+目的地区域匹配（印尼/马来/泰国/越南/菲律宾/新加坡/美国）
5. 品类克重库：鸡蛋盒780g/防晒帽150g/脏衣篮1.1kg/收纳盒350g/通用300g；--weight 可覆盖

用法示例：
  python ecom_report_tool.py --product "图.jpg" --purchase-image "1688.png" --platform-image "shopee.png" \
      --product-name "产品名" --platform shopee-id --logistics sea \
      --purchase 8.50 --benchmark 95000 --weight 780 --currency IDR --search-json "data.json"
"""
import argparse
import datetime
import hashlib
import json
import os
import re
import sqlite3
import sys

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 1. 平台配置（可扩展）
# ============================================================
PLATFORMS = {
    "shopee-id": {
        "name": "Shopee 印尼本土店(PT)", "currency": "IDR", "region": "indonesia",
        "commission": 0.05, "trans": 0.02, "pay": 0.01, "cert": [],
    },
    "shopee-id-cb": {
        "name": "Shopee 印尼跨境店(CB)", "currency": "IDR", "region": "indonesia",
        "commission": 0.08, "trans": 0.06, "pay": 0.01, "cert": [],
    },
    "lazada-id": {
        "name": "Lazada 印尼(LGS)", "currency": "IDR", "region": "indonesia",
        "commission": 0.06, "trans": 0.02, "pay": 0.01, "cert": [],
    },
    "tiktok-id": {
        "name": "TikTok Shop 印尼", "currency": "IDR", "region": "indonesia",
        "commission": 0.02, "trans": 0.02, "pay": 0.01, "cert": [],
    },
    "shopee-my": {
        "name": "Shopee 马来(本地店)", "currency": "MYR", "region": "malaysia",
        "commission": 0.05, "trans": 0.02, "pay": 0.01, "cert": [],
    },
    "lazada-my": {
        "name": "Lazada 马来", "currency": "MYR", "region": "malaysia",
        "commission": 0.06, "trans": 0.02, "pay": 0.01, "cert": [],
    },
    "shopee-th": {
        "name": "Shopee 泰国(本地店)", "currency": "THB", "region": "thailand",
        "commission": 0.05, "trans": 0.02, "pay": 0.01, "cert": [],
    },
    "shopee-vn": {
        "name": "Shopee 越南(本地店)", "currency": "VND", "region": "vietnam",
        "commission": 0.05, "trans": 0.02, "pay": 0.01, "cert": [],
    },
    "shopee-ph": {
        "name": "Shopee 菲律宾(本地店)", "currency": "PHP", "region": "philippines",
        "commission": 0.05, "trans": 0.02, "pay": 0.01, "cert": [],
    },
    "shopee-sg": {
        "name": "Shopee 新加坡(本地店)", "currency": "SGD", "region": "singapore",
        "commission": 0.05, "trans": 0.02, "pay": 0.01, "cert": [],
    },
    "amazon-us": {
        "name": "Amazon 美国站", "currency": "USD", "region": "us",
        "commission": 0.15, "trans": 0.01, "pay": 0.01, "cert": [],
    },
}
# 区域物流费率库（国际段海运 ¥/kg；本地段：轻小件≤500g 按件价，>500g 按 ¥/kg×克重）
REGIONS = {
    "indonesia":   {"name": "印尼",   "sea_per_kg": 3.0, "local_small": 2.0, "local_per_kg": 4.0, "local_note": "JNE/J&T"},
    "malaysia":    {"name": "马来",   "sea_per_kg": 3.0, "local_small": 2.0, "local_per_kg": 2.5, "local_note": "Pos/J&T"},
    "thailand":    {"name": "泰国",   "sea_per_kg": 3.5, "local_small": 2.0, "local_per_kg": 3.0, "local_note": "Flash Express"},
    "vietnam":     {"name": "越南",   "sea_per_kg": 3.5, "local_small": 2.0, "local_per_kg": 3.0, "local_note": "GHN/J&T"},
    "philippines": {"name": "菲律宾", "sea_per_kg": 3.5, "local_small": 2.0, "local_per_kg": 3.0, "local_note": "J&T"},
    "singapore":   {"name": "新加坡", "sea_per_kg": 4.0, "local_small": 2.5, "local_per_kg": 3.0, "local_note": "J&T"},
    "us":          {"name": "美国",   "sea_per_kg": 8.0, "local_small": 15.0, "local_per_kg": 15.0, "local_note": "USPS/FedEx"},
}
# 币种汇率（1 当地货币 = ? CNY）
CURRENCIES = {
    "IDR": 0.0003782, "USD": 7.20, "MYR": 1.65, "THB": 0.21,
    "VND": 0.00029, "PHP": 0.13, "SGD": 5.50, "CNY": 1.0,
}
# 品类克重库（g/件）；--weight 可覆盖
CATEGORY_WEIGHT = [
    (r"鸡蛋|蛋盒|egg|蛋托", 780),
    (r"帽|cap|hat|防晒", 150),
    (r"脏衣篮|收纳篮|洗衣篮|laundry|basket|hamper", 1100),
    (r"收纳盒|收纳箱|storage|box|盒", 350),
]
DEFAULT_WEIGHT = 300
# 物流方案（v3：无空运；mixed=首轮海运+海运补货）
LOGISTICS = {
    "sea":   {"label": "海运 LCL（20-35天）", "days": "20-35天"},
    "mixed": {"label": "海运+海运补货（首轮大批量+补货）", "days": "全程海运20-35天"},
}
# 强制资质规则
CERT_RULES = [
    ("*", r"food\s*contact|食品接触|餐盒|奶瓶|水杯|吸管", "SNI 7323:2008（食品接触塑料强制）"),
    ("*", r"cosmetic|化妆品|面膜|彩妆|口红|防晒霜|防晒乳|防晒喷雾|sunscreen|sunblock|隔离霜|粉底|睫毛膏", "BPOM 注册（化妆品强制）"),
    ("*", r"drug|药品|保健品|维生素|protein|蛋白粉", "BPOM 注册（药品/保健品强制）"),
    ("*", r"electronic|电子|充电|电源|灯|light|battery|电池|耳机|phone|手机", "SNI 强制认证（电子产品）"),
    ("*", r"baby|儿童|玩具|toy|child", "SNI 强制认证（儿童用品）"),
    ("*", r"vape|电子烟|alcohol|酒|香烟|cigarette", "平台禁售/限制类目"),
    ("shopee-id|shopee-id-cb|lazada-id|tiktok-id", r"halal|清真", "Halal 认证（食品类强制）"),
]
# 内置平台价格库（auto-search 模式）
AUTO_PRICE_LIB = {
    "laundry|脏衣篮|收纳篮|脏衣篓|basket|hamper": {
        "comp1688": [("无品牌普料壁挂脏衣篮", 3.60, "塑料薄"), ("可折叠脏衣篓(热销)", 5.50, "价格战最烈"), ("带盖壁挂脏衣篮", 10.50, "高价小众")],
        "compPlat": [("无品牌塑料脏衣篮", 45000, "红海带底部"), ("入门折叠布艺款", 65000, "主销带"), ("大众织物款", 85000, "大众段")],
    },
    "鸡蛋|蛋盒|蛋托|egg": {
        "comp1688": [("双层32格PET鸡蛋盒(义乌购同款)", 8.50, "起购200件/箱20个"), ("双层32格白色", 8.50, "库存11832"), ("双层32格灰色", 8.50, "库存13961"), ("双层40格白色", 15.00, "库存10070"), ("源头工厂款", 2.88, "台州沐鸣超级工厂"), ("源头工厂款", 3.93, "义乌创购5年店")],
        "compPlat": [("OneFamily 本品FLASH", 45540, "4.8分/10W+售"), ("30蛋3层PET收纳盒", 50000, "印尼在售"), ("基础款鸡蛋盒", 35000, "红海底部"), ("双层抽屉款", 60000, "主销带")],
    },
}
DEFAULT_LIB = {"comp1688": [("同类1688款(未匹配内置库)", 5.00, "请补充")], "compPlat": [("同类平台款", 50000, "请补充")]}

# ============================================================
# 2. 记忆库
# ============================================================
MEMORY_DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ecom_report_memory.db")


def init_memory(db_path=MEMORY_DB):
    conn = sqlite3.connect(db_path)
    conn.execute("""CREATE TABLE IF NOT EXISTS reports(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_hash TEXT, product_name TEXT, platform TEXT, logistics TEXT,
        purchase REAL, benchmark REAL, result TEXT, note TEXT,
        file_path TEXT, created_at TEXT)""")
    conn.commit()
    return conn


def product_hash(name):
    return hashlib.md5(name.strip().lower().encode("utf-8")).hexdigest()[:16]


def check_duplicate(conn, product_name, platform):
    cur = conn.execute(
        "SELECT id, product_name, platform, result, created_at, file_path FROM reports "
        "WHERE product_hash=? AND platform=? ORDER BY id DESC LIMIT 1",
        (product_hash(product_name), platform))
    return cur.fetchone()


def save_report(conn, product_name, platform, logistics, purchase, benchmark, result, note, file_path):
    conn.execute(
        "INSERT INTO reports(product_hash, product_name, platform, logistics, purchase, benchmark, result, note, file_path, created_at) "
        "VALUES(?,?,?,?,?,?,?,?,?,?)",
        (product_hash(product_name), product_name, platform, logistics, purchase, benchmark, result, note,
         file_path, datetime.datetime.now().isoformat(timespec="seconds")))
    conn.commit()

# ============================================================
# 3. 工具函数
# ============================================================
def fmt_cny(v): return f"¥{v:,.2f}"
def fmt_pct(v): return f"{v*100:.1f}%"
def num(v): 
    try: return float(v)
    except: return 0.0

def fmt_money(amount, cur="IDR"):
    if cur == "IDR": return f"Rp{amount:,.0f}"
    if cur == "USD": return f"${amount:,.2f}"
    if cur == "MYR": return f"RM{amount:,.2f}"
    if cur == "THB": return f"฿{amount:,.0f}"
    if cur == "VND": return f"₫{amount:,.0f}"
    if cur == "PHP": return f"₱{amount:,.2f}"
    if cur == "SGD": return f"S${amount:,.2f}"
    return f"{amount:,.2f}"

def to_cny(amount, cur):
    return num(amount) * CURRENCIES.get(cur, 1.0)

def pick_weight(product_name, weight):
    """返回单件克重(g)：--weight 优先，否则按品类库估算"""
    if weight and weight > 0:
        return weight
    for pattern, w in CATEGORY_WEIGHT:
        if re.search(pattern, product_name, re.I):
            return w
    return DEFAULT_WEIGHT

def region_fee(region_key, weight_g):
    """按克重算运费：国际段 + 本地段（CNY/件）"""
    r = REGIONS[region_key]
    kg = weight_g / 1000.0
    sea = r["sea_per_kg"] * kg                       # 国际段海运
    if weight_g <= 500:
        local = r["local_small"]                     # 轻小件按件
    else:
        local = r["local_per_kg"] * kg               # 重件按 kg
    return sea, local

def calc_profit(price_cur, ad_rate, purchase_cny, plat, weight_g, extra_cny=0.0):
    """单件净利 CNY（v3：按克重算运费）"""
    rev_cny = to_cny(price_cur, plat["currency"])
    sea_fee, local_fee = region_fee(plat["region"], weight_g)
    cost = purchase_cny + 0.5 + sea_fee + 2.0 + local_fee + 1.0
    fee = rev_cny * (plat["commission"] + plat["trans"] + plat["pay"] + ad_rate)
    return rev_cny - fee - cost

def calc_rate(price_cur, ad_rate, purchase_cny, plat, weight_g, extra_cny=0.0):
    rev = to_cny(price_cur, plat["currency"])
    return 0.0 if rev == 0 else calc_profit(price_cur, ad_rate, purchase_cny, plat, weight_g, extra_cny) / rev

# ============================================================
# 4. 样式
# ============================================================
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
SUB_FONT = Font(name="微软雅黑", size=10, color="595959")
BODY_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
GREEN_FONT = Font(name="微软雅黑", size=10, bold=True, color="1E7B34")
RED_FONT = Font(name="微软雅黑", size=10, bold=True, color="C00000")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
BAD_FILL = PatternFill("solid", fgColor="FCE4EC")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

# ============================================================
# 5. 资质核查
# ============================================================
def check_certification(product_name, platform_key):
    needed = []
    for plat_wild, pattern, cert_name in CERT_RULES:
        wilds = plat_wild.split("|")
        if "*" not in wilds and platform_key not in wilds:
            continue
        if re.search(pattern, product_name, re.I):
            needed.append(cert_name)
    return (len(needed) > 0, needed)

# ============================================================
# 6. 报表生成
# ============================================================
def embed_images(ws, anchor, img_paths, col_span=4, img_h=110):
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    col_str, row_str = coordinate_from_string(anchor)
    c_start = column_index_from_string(col_str)
    row = int(row_str)
    c = c_start
    for label, path in img_paths.items():
        if not path or not os.path.exists(path):
            ws.cell(row=row, column=c, value=f"[{label}:缺图]").font = SUB_FONT
        else:
            try:
                img = XLImage(path)
                img.height = img_h
                img.width = img_h
                ws.add_image(img, f"{get_column_letter(c)}{row}")
            except Exception:
                ws.cell(row=row, column=c, value=f"[{label}:图读失败]").font = SUB_FONT
        c += 1
        if c >= c_start + col_span:
            break


def build_report(args, plat, logi, purchase_cny, benchmark, ad_rates, discount_steps,
                 comp_1688, comp_plat, weight_g, result_meta):
    out_path = args.output
    wb = Workbook()
    ws = wb.active
    ws.title = "产品与竞品图"
    ncol = 3 + 4 * 2
    region = REGIONS[plat["region"]]
    sea_fee, local_fee = region_fee(plat["region"], weight_g)

    # ---------- Sheet 0: 产品与竞品图 ----------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws["A1"] = f"产品与竞品图对照 — {args.product_name}（{plat['name']} / {logi['label']}）"
    ws["A1"].font = TITLE_FONT
    img_headers = ["产品图", "1688采购图", "平台售卖图", "说明/备注"]
    for i, h in enumerate(img_headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER
    embed_images(ws, "A4", {"产品图": args.product_image, "1688采购图": args.purchase_image, "平台售卖图": args.platform_image})
    ws.cell(row=4, column=4, value=(
        f"产品：{args.product_name}\n平台：{plat['name']}\n物流：{logi['label']}\n"
        f"采购价：{fmt_cny(purchase_cny)}\n克重：{weight_g}g/件\n"
        f"基准售价：{fmt_money(benchmark, plat['currency'])} ≈ {fmt_cny(to_cny(benchmark, plat['currency']))}\n"
        f"国际段海运：{fmt_cny(sea_fee)}/件 | 本地段({region['local_note']})：{fmt_cny(local_fee)}/件"
    )).font = BODY_FONT
    ws.cell(row=4, column=4).alignment = LEFT
    for cc in range(1, 5):
        ws.cell(row=4, column=cc).border = BORDER
    ws.row_dimensions[4].height = 110
    for i, w in enumerate([16, 16, 16, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    # ---------- Sheet 1: 利润明细 ----------
    ws1 = wb.create_sheet("利润明细")
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws1["A1"] = f"利润明细 — {args.product_name} × {plat['name']}"
    ws1["A1"].font = TITLE_FONT
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws1["A2"] = (f"基准售价 {fmt_money(benchmark, plat['currency'])}（≈{fmt_cny(to_cny(benchmark, plat['currency']))}） | "
                 f"采购价 {fmt_cny(purchase_cny)} | 克重 {weight_g}g | 国际段海运 {fmt_cny(sea_fee)}/件 + 本地段 {fmt_cny(local_fee)}/件 | "
                 f"佣金{plat['commission']*100:.0f}%+交易{plat['trans']*100:.0f}%+支付{plat['pay']*100:.0f}%")
    ws1["A2"].font = SUB_FONT
    headers = ["降价档", f"售价({plat['currency']})", "售价(CNY)"]
    for ad in ad_rates:
        headers += [f"广告{int(ad*100)}%净利", "利润率"]
    row = 3
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER
    row += 1
    for d in discount_steps:
        price = int(benchmark * (1 - d))
        label = "原价" if d == 0 else f"降{d*100:.0f}%"
        ws1.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws1.cell(row=row, column=2, value=fmt_money(price, plat["currency"])).font = BODY_FONT
        ws1.cell(row=row, column=3, value=f"≈{fmt_cny(to_cny(price, plat['currency']))}").font = BODY_FONT
        col = 4
        for ad in ad_rates:
            p = calc_profit(price, ad, purchase_cny, plat, weight_g)
            r = calc_rate(price, ad, purchase_cny, plat, weight_g)
            c1 = ws1.cell(row=row, column=col, value=fmt_cny(p)); c1.alignment = RIGHT; col += 1
            c2 = ws1.cell(row=row, column=col, value=fmt_pct(r)); c2.alignment = RIGHT; col += 1
            for c in (c1, c2):
                c.border = BORDER
                c.font = GREEN_FONT if p > 0 else RED_FONT
        for cc in range(1, len(headers) + 1):
            ws1.cell(row=row, column=cc).border = BORDER
        row += 1
    ws1.freeze_panes = "A4"
    widths = [16, 12, 12] + [11, 9] * len(ad_rates)
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ---------- Sheet 2: 竞品全景 ----------
    ws2 = wb.create_sheet("竞品全景")
    ws2.merge_cells("A1:G1")
    ws2["A1"] = "竞品全景 — 1688 供应链价格 vs 平台售价（来源：AI 实时搜索 / 用户截图 / 内置库）"
    ws2["A1"].font = TITLE_FONT
    headers2 = ["序号", "商品/卖家", "1688价(CNY)", f"平台价({plat['currency']})", "折合CNY", "月销/热度", "对标痛点"]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER
    r = 3
    ws2.cell(row=r, column=1, value="本品基准").font = BOLD_FONT
    ws2.cell(row=r, column=2, value=args.product_name).font = BODY_FONT
    ws2.cell(row=r, column=3, value=fmt_cny(purchase_cny)).font = BODY_FONT
    ws2.cell(row=r, column=4, value=fmt_money(benchmark, plat["currency"])).font = BODY_FONT
    ws2.cell(row=r, column=5, value=f"≈{fmt_cny(to_cny(benchmark, plat['currency']))}").font = BODY_FONT
    ws2.cell(row=r, column=6, value="★ 待售").font = BODY_FONT
    ws2.cell(row=r, column=7, value="基准锚点").font = BODY_FONT
    for cc in range(1, 8):
        ws2.cell(row=r, column=cc).border = BORDER; ws2.cell(row=r, column=cc).fill = WARN_FILL
    r += 1
    for i, (name, price, pain) in enumerate(comp_1688, 1):
        ws2.cell(row=r, column=1, value=f"1688-{i}").font = BODY_FONT
        ws2.cell(row=r, column=2, value=name).font = BODY_FONT
        ws2.cell(row=r, column=3, value=fmt_cny(price)).font = BODY_FONT
        ws2.cell(row=r, column=4, value="—").font = BODY_FONT
        ws2.cell(row=r, column=5, value="—").font = BODY_FONT
        ws2.cell(row=r, column=6, value="—").font = BODY_FONT
        ws2.cell(row=r, column=7, value=pain).font = BODY_FONT
        for cc in range(1, 8):
            ws2.cell(row=r, column=cc).border = BORDER; ws2.cell(row=r, column=cc).alignment = LEFT
        r += 1
    for i, (name, price, note) in enumerate(comp_plat, 1):
        ws2.cell(row=r, column=1, value=f"平台-{i}").font = BODY_FONT
        ws2.cell(row=r, column=2, value=name).font = BODY_FONT
        ws2.cell(row=r, column=3, value="—").font = BODY_FONT
        ws2.cell(row=r, column=4, value=fmt_money(price, plat["currency"])).font = BODY_FONT
        ws2.cell(row=r, column=5, value=f"≈{fmt_cny(to_cny(price, plat['currency']))}").font = BODY_FONT
        ws2.cell(row=r, column=6, value=note).font = BODY_FONT
        ws2.cell(row=r, column=7, value="").font = BODY_FONT
        for cc in range(1, 8):
            ws2.cell(row=r, column=cc).border = BORDER; ws2.cell(row=r, column=cc).alignment = LEFT
        r += 1
    for i, w in enumerate([10, 36, 13, 16, 12, 18, 34], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A3"

    # ---------- Sheet 3: 决策结论 ----------
    ws3 = wb.create_sheet("决策结论")
    ws3.merge_cells("A1:B1")
    ws3["A1"] = "决策结论 — 记忆库/资质/利润三重检查"
    ws3["A1"].font = TITLE_FONT
    ws3.cell(row=2, column=1, value="检查项").font = HEADER_FONT
    ws3.cell(row=2, column=2, value="结果").font = HEADER_FONT
    for c in (1, 2):
        ws3.cell(row=2, column=c).fill = HEADER_FILL; ws3.cell(row=2, column=c).alignment = CENTER
    r = 3
    for name, ok, detail in result_meta.get("checks", []):
        ws3.cell(row=r, column=1, value=name).font = BOLD_FONT
        ws3.cell(row=r, column=2, value=detail).font = GREEN_FONT if ok else RED_FONT
        ws3.cell(row=r, column=1).border = BORDER; ws3.cell(row=r, column=2).border = BORDER
        ws3.cell(row=r, column=1).fill = OK_FILL if ok else BAD_FILL
        ws3.cell(row=r, column=2).fill = OK_FILL if ok else BAD_FILL
        r += 1
    r += 1
    ws3.cell(row=r, column=1, value="定价建议").font = BOLD_FONT
    r += 1
    for a in result_meta.get("advice", []):
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws3.cell(row=r, column=1, value=a).font = BODY_FONT
        ws3.cell(row=r, column=1).alignment = LEFT
        r += 1
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 110

    # ---------- Sheet 4: 供应链物流（v3：仅海运，按克重） ----------
    ws4 = wb.create_sheet("供应链物流")
    ws4.merge_cells("A1:G1")
    ws4["A1"] = f"供应链与物流（按克重计费） — {logi['label']} | 单件克重 {weight_g}g | 区域：{region['name']}（{region['local_note']}）"
    ws4["A1"].font = TITLE_FONT
    headers4 = ["环节", "说明", "计费方式", "费用(CNY/件)", "备注", "", ""]
    for c, h in enumerate(headers4, 1):
        cell = ws4.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER
    rows4 = [
        ("① 采购", "工厂采购", "按件", purchase_cny, f"采购价 {fmt_cny(purchase_cny)}", "", ""),
        ("② 国内集运", "工厂→集货仓", "按件", 0.50, "含国内运输", "", ""),
        ("③ 国际段海运", "中国→" + region["name"], f"{region['sea_per_kg']}¥/kg × {weight_g}g", sea_fee, "海运LCL 20-35天", "", ""),
        ("④ 清关+仓", "清关入库", "按件", 2.00, "需进口资质", "", ""),
        ("⑤ 本地配送", "仓→买家", ("≤500g按件" if weight_g <= 500 else f"{region['local_per_kg']}¥/kg × {weight_g}g"), local_fee, f"{region['local_note']} 1-3天", "", ""),
        ("⑥ 包装", "彩盒+袋", "按件", 1.00, "防压", "", ""),
    ]
    r = 3
    for row_ in rows4:
        for c, v in enumerate(row_, 1):
            cell = ws4.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT; cell.border = BORDER; cell.alignment = LEFT
        r += 1
    r += 1
    total = purchase_cny + 0.5 + sea_fee + 2.0 + local_fee + 1.0
    ws4.cell(row=r, column=1, value="单件到岸成本").font = BOLD_FONT
    ws4.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws4.cell(row=r, column=2, value=f"{fmt_cny(purchase_cny)}+0.5+{fmt_cny(sea_fee)}+2+{fmt_cny(local_fee)}+1 = {fmt_cny(total)}").font = BOLD_FONT
    ws4.cell(row=r, column=5, value="全海运（无空运方案）").font = RED_FONT
    for cc in range(1, 8):
        ws4.cell(row=r, column=cc).border = BORDER; ws4.cell(row=r, column=cc).fill = WARN_FILL
    for i, w in enumerate([16, 20, 22, 16, 30, 8, 8], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = "A3"

    wb.save(out_path)
    return out_path


# ============================================================
# 7. 主流程
# ============================================================
def resolve_images(args):
    if args.product_image:
        return args.product_image, args.purchase_image, args.platform_image
    if args.image_dir and os.path.isdir(args.image_dir):
        files = os.listdir(args.image_dir)
        def find(patterns, exclude=()):
            for f in sorted(files):
                low = f.lower()
                if any(p in low for p in patterns) and not any(e in low for e in exclude):
                    return os.path.join(args.image_dir, f)
            return None
        product = find(["product", "产品", "主图", "jpg", "jpeg", "png"], exclude=["1688", "shopee", "竞品", "clipboard"])
        purchase = find(["1688", "采购"])
        platform = find(["shopee", "lazada", "tiktok", "amazon", "售卖", "平台"])
        return product, purchase, platform
    return None, None, None


def pick_auto_competitors(product_name):
    for key, lib in AUTO_PRICE_LIB.items():
        if re.search(key, product_name, re.I):
            return lib["comp1688"], lib["compPlat"]
    return DEFAULT_LIB["comp1688"], DEFAULT_LIB["compPlat"]


def main(argv=None):
    ap = argparse.ArgumentParser(description="电商产品图片分析报表工具 v3.0")
    ap.add_argument("--config", help="JSON 配置文件")
    ap.add_argument("--product", dest="product_image", help="产品图路径")
    ap.add_argument("--purchase-image", dest="purchase_image", help="1688采购图路径")
    ap.add_argument("--platform-image", dest="platform_image", help="平台售卖图路径")
    ap.add_argument("--image-dir", dest="image_dir", help="图片目录")
    ap.add_argument("--product-name", dest="product_name", default="", help="产品名称")
    ap.add_argument("--platform", default="shopee-id", choices=list(PLATFORMS.keys()), help="售卖平台")
    ap.add_argument("--logistics", default="sea", choices=list(LOGISTICS.keys()), help="物流方式：sea/mixed（v3无空运）")
    ap.add_argument("--purchase", type=float, default=0.0, help="采购价 CNY")
    ap.add_argument("--benchmark", type=float, default=0.0, help="基准售价（当地货币）")
    ap.add_argument("--weight", type=float, default=0.0, help="单件克重 g（默认按品类库估算）")
    ap.add_argument("--currency", default="", help="币种（默认跟随平台；可自主选 IDR/USD/MYR/THB/VND/PHP/SGD/CNY）")
    ap.add_argument("--ad", default="0,5,10,15", help="广告率百分比，逗号分隔")
    ap.add_argument("--discount", default="5,10,15,20,25", help="降价阶梯百分比（不含原价）")
    ap.add_argument("--extra-cost", dest="extra_cost", type=float, default=0.0, help="其他固定成本 CNY/件")
    ap.add_argument("--search-json", dest="search_json", help="AI 搜索竞品数据 JSON")
    ap.add_argument("--auto-search", action="store_true", help="内置价格库匹配")
    ap.add_argument("--output", default="", help="输出 xlsx 路径")
    ap.add_argument("--force", action="store_true", help="跳过记忆库去重")
    args = ap.parse_args(argv)

    if args.config and os.path.exists(args.config):
        with open(args.config, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        for k, v in cfg.items():
            if getattr(args, k, None) in (None, "", 0.0, False, []) or k in ("output", "product_name"):
                setattr(args, k, v)

    plat = PLATFORMS[args.platform]
    logi = LOGISTICS[args.logistics]
    if not args.currency:
        args.currency = plat["currency"]
    ad_rates = [float(x) / 100 for x in args.ad.split(",")]
    discount_steps = [float(x) / 100 for x in args.discount.split(",")]

    prod_img, pur_img, plat_img = resolve_images(args)
    if not args.product_name:
        base = os.path.basename(prod_img) if prod_img else "未命名产品"
        args.product_name = os.path.splitext(base)[0]

    # 克重：--weight 优先，否则品类估算
    weight_g = pick_weight(args.product_name, args.weight)
    purchase = args.purchase if args.purchase > 0 else 4.0

    # ---------- 禁用条件 ----------
    checks = []
    need_cert, cert_list = check_certification(args.product_name, args.platform)
    checks.append(("资质检查", not need_cert,
                   "✅ 无强制资质门槛" if not need_cert else f"❌ 需强制资质：{'、'.join(cert_list)}（停止出表）"))
    if need_cert:
        result_meta = {"checks": checks, "advice": [f"该产品触发强制资质（{'、'.join(cert_list)}），按禁用规则停止生成报表。"]}
        _emit_blocked(args, result_meta, "资质不足", cert_list)
        return 1

    conn = init_memory()
    dup = None if args.force else check_duplicate(conn, args.product_name, args.platform)
    if dup:
        checks.append(("记忆库去重", False,
                       f"❌ 重复使用：'{dup[1]}' 曾在 {dup[2]} 生成过（{dup[3]}，{dup[4]}）→ {dup[5]}（停止出表）"))
        result_meta = {"checks": checks, "advice": [f"该产品在 {args.platform} 已分析过。如需重新分析请加 --force。"]}
        _emit_blocked(args, result_meta, "重复使用", dup[5])
        conn.close()
        return 1
    checks.append(("记忆库去重", True, "✅ 新项目，无重复记录"))

    comp_1688, comp_plat = [], []
    if args.search_json and os.path.exists(args.search_json):
        with open(args.search_json, "r", encoding="utf-8") as f:
            data = json.load(f)
        comp_1688 = data.get("competitors_1688", [])
        comp_plat = data.get("competitors_platform", [])
        checks.append(("自动搜索", True, f"✅ 已注入 AI 实时搜索结果（1688 {len(comp_1688)} 款 / 平台 {len(comp_plat)} 款）"))
    elif args.auto_search:
        comp_1688, comp_plat = pick_auto_competitors(args.product_name)
        checks.append(("自动搜索", True, f"✅ 内置价格库匹配（1688 {len(comp_1688)} 款 / 平台 {len(comp_plat)} 款）"))
    else:
        checks.append(("自动搜索", True, "ℹ️ 未启用自动搜索"))

    main_ad = ad_rates[1] if len(ad_rates) > 1 else ad_rates[0]
    main_p = calc_profit(args.benchmark, main_ad, purchase, plat, weight_g, args.extra_cost)
    main_rate = calc_rate(args.benchmark, main_ad, purchase, plat, weight_g, args.extra_cost)
    checks.append(("利润判定", main_p > 0,
                   f"{'✅' if main_p > 0 else '❌'} 主力场景（基准原价+{int(main_ad*100)}%广告+{args.logistics}）"
                   f"单件净利 {fmt_cny(main_p)}，利润率 {fmt_pct(main_rate)}"))
    if main_p <= 0:
        result_meta = {"checks": checks, "advice": [f"主力场景利润 {fmt_cny(main_p)} ≤ 0，按禁用规则停止生成报表。建议：谈判采购价 / 换物流 / 提高售价。"]}
        _emit_blocked(args, result_meta, "负利润", fmt_cny(main_p))
        conn.close()
        return 1

    region = REGIONS[plat["region"]]
    sea_fee, local_fee = region_fee(plat["region"], weight_g)
    advice = [
        f"✅ 项目可行：主力场景（基准原价+{int(main_ad*100)}%广告+{args.logistics}）单件净利 {fmt_cny(main_p)}，利润率 {fmt_pct(main_rate)}",
        f"物流（v3 全海运）：国际段 {fmt_cny(sea_fee)}/件（{region['sea_per_kg']}¥/kg × {weight_g}g）+ 本地段 {region['local_note']} {fmt_cny(local_fee)}/件；克重 {weight_g}g 由品类库估算（--weight 可覆盖）",
        f"币种：{args.currency}（1 {args.currency} = {CURRENCIES[args.currency]} CNY），全部金额已换算人民币",
        "定价建议：日常价建议落在基准价 60-85% 区间保利润；FLASH 促销价若亏本仅作引流款。",
    ]
    result_meta = {"checks": checks, "advice": advice}
    out_path = args.output or os.path.join(
        os.path.expanduser("~"), "Desktop", f"{args.product_name}_{args.platform}_选品分析报表.xlsx")
    try:
        built = build_report(args, plat, logi, purchase, args.benchmark,
                             ad_rates, discount_steps, comp_1688, comp_plat, weight_g, result_meta)
    except Exception as e:
        import traceback
        traceback.print_exc()
        result_meta = {"checks": checks, "advice": [f"报表生成失败：{e}"]}
        _emit_blocked(args, result_meta, "生成失败", str(e))
        conn.close()
        return 1

    save_report(conn, args.product_name, args.platform, args.logistics, purchase,
                args.benchmark, "OK", f"weight={weight_g}g,cur={args.currency}", out_path)
    conn.close()

    result = {
        "status": "OK", "product": args.product_name, "platform": args.platform,
        "logistics": args.logistics, "weight_g": weight_g, "currency": args.currency,
        "output": built, "main_profit": main_p, "main_rate": main_rate,
        "checks": [{"name": n, "ok": b, "detail": d} for n, b, d in checks],
    }
    result_json = os.path.join(os.path.dirname(out_path), f"{args.product_name}_result.json")
    with open(result_json, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print("DONE")
    print("OUTPUT:", built)
    print("WEIGHT_G:", weight_g, "CURRENCY:", args.currency)
    print("MAIN_PROFIT:", fmt_cny(main_p), "RATE:", fmt_pct(main_rate))
    print("CHECKS:", json.dumps([n for n, b, d in checks], ensure_ascii=False))
    return 0


def _emit_blocked(args, result_meta, reason, detail):
    out_dir = os.path.expanduser("~") + os.sep + "Desktop"
    block_file = os.path.join(out_dir, f"{args.product_name}_停止原因报告.txt")
    with open(block_file, "w", encoding="utf-8") as f:
        f.write(f"产品：{args.product_name}\n平台：{args.platform}\n物流：{args.logistics}\n\n")
        f.write(f"【停止原因】{reason}: {detail}\n\n检查明细：\n")
        for name, ok, d in result_meta["checks"]:
            f.write(f"- {name}: {d}\n")
        f.write("\n建议：\n")
        for a in result_meta["advice"]:
            f.write(f"- {a}\n")
    print("BLOCKED:", reason, "-", detail)
    print("BLOCKED_REPORT:", block_file)
    print("NO_REPORT_GENERATED")


if __name__ == "__main__":
    sys.exit(main())
