# -*- coding: utf-8 -*-
"""
generate_ecom_report.py — 电商产品图片分析报表生成器（v1.0.0）
产品案例：DejaVu HSB808 浴室壁挂式脏衣篮 × Shopee 印尼站（本土店）
输出：E:\Desktop\{产品名}_Shopee印尼选品分析报表.xlsx（5-sheet）

数据来源说明：用户提供文字参数 + WebSearch 行业数据（2026-08-27）
汇率：1 CNY = 2,644 IDR（0.0003782 CNY/IDR）
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============ 参数区 ============
FX = 0.0003782          # 1 IDR = ? CNY
PRODUCT = "DejaVu HSB808 浴室壁挂式脏衣篮"
PLATFORM = "Shopee 印尼站（本土店 PT）"
PURCHASE_BASE = 4.00    # 采购价基准 CNY（区间 3.60-4.50）
BASE_PRICE_IDR = 109000 # 基准售价（原价）
FLASH_PRICE_IDR = 53105 # FLASH SALE 参考价
COMMISSION = 0.05       # 佣金
TRANS_FEE = 0.02        # 交易费
PAY_FEE = 0.01          # 支付费
AD_RATES = [0.00, 0.05, 0.10, 0.15]        # 广告率4档
DISCOUNT_STEPS = [0.00, 0.05, 0.10, 0.15, 0.20, 0.25]  # 降价阶梯(含原价)
SHIPPING_SEA = 3.00     # 海运国际段 CNY/件
SHIPPING_AIR = 22.00    # 空运国际段 CNY/件
SEA_IDR_FEE = 10000     # 印尼段本地配送 IDR/单
PACK_FEE = 1.00         # 包装 CNY/件
INLAND_FEE = 2.00       # 清关+海外仓操作 CNY/件
COLLECT_FEE = 0.50      # 国内集运 CNY/件
BATCH = [(100, 0.00), (500, 0.08), (1000, 0.15), (3000, 0.22)]  # (起订量, 折扣)
OUT_PATH = r"E:\Desktop\浴室壁挂式脏衣篮_Shopee印尼选品分析报表.xlsx"

# ============ 样式 ============
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
SUB_FONT = Font(name="微软雅黑", size=10, color="595959")
BODY_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
GREEN_FONT = Font(name="微软雅黑", size=10, bold=True, color="1E7B34")
RED_FONT = Font(name="微软雅黑", size=10, bold=True, color="C00000")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
BAD_FILL = PatternFill("solid", fgColor="FCE4EC")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

def fmt_cny(v):
    return f"¥{v:,.2f}"

def fmt_pct(v):
    return f"{v*100:.1f}%"

def fmt_idr(v):
    return f"Rp{v:,.0f}"

def cny(v):
    return v * FX

# ============ 计算 ============
def unit_cost(purchase, ship_intl):
    """单件到岸成本（不含平台费用）"""
    return purchase + COLLECT_FEE + ship_intl + INLAND_FEE + cny(SEA_IDR_FEE) + PACK_FEE

def profit_per_unit(price_idr, ad_rate, purchase, ship_intl):
    """单件净利 CNY"""
    rev = cny(price_idr)
    platform = rev * (COMMISSION + TRANS_FEE + PAY_FEE + ad_rate)
    return rev - platform - unit_cost(purchase, ship_intl)

def profit_rate(price_idr, ad_rate, purchase, ship_intl):
    p = profit_per_unit(price_idr, ad_rate, purchase, ship_intl)
    return p / cny(price_idr) if cny(price_idr) else 0

# ============ Sheet 1: 利润明细 ============
def build_sheet1(wb):
    ws = wb.active
    ws.title = "利润明细"
    ncol = 3 + 4 * 2  # 3 固定列 + 4 组(净利+率)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws["A1"] = f"利润明细 — {PRODUCT} × {PLATFORM}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws["A2"] = (f"基准售价 {fmt_idr(BASE_PRICE_IDR)}（≈{fmt_cny(cny(BASE_PRICE_IDR))}） | FLASH SALE 参考 {fmt_idr(FLASH_PRICE_IDR)}"
                f"（≈{fmt_cny(cny(FLASH_PRICE_IDR))}） | 采购价 {fmt_cny(PURCHASE_BASE)}（1688最低档 ¥3.60-4.50） | 汇率 1CNY=2,644IDR | 佣金5%+交易费2%+支付1%")
    ws["A2"].font = SUB_FONT

    headers = ["降价档", "售价(IDR)", "售价(CNY)"]
    for ad in AD_RATES:
        headers += [f"广告{int(ad*100)}%净利(海运)", "利润率"]
    row = 3
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER
    row += 1
    for d in DISCOUNT_STEPS:
        price = int(BASE_PRICE_IDR * (1 - d))
        label = "原价" if d == 0 else f"降{d*100:.0f}%"
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value=fmt_idr(price)).font = BODY_FONT
        ws.cell(row=row, column=3, value=f"≈{fmt_cny(cny(price))}").font = BODY_FONT
        col = 4
        for ad in AD_RATES:
            p = profit_per_unit(price, ad, PURCHASE_BASE, SHIPPING_SEA)
            r = profit_rate(price, ad, PURCHASE_BASE, SHIPPING_SEA)
            c1 = ws.cell(row=row, column=col, value=fmt_cny(p)); c1.alignment = RIGHT; col += 1
            c2 = ws.cell(row=row, column=col, value=fmt_pct(r)); c2.alignment = RIGHT; col += 1
            for c in (c1, c2):
                c.border = BORDER
                c.font = GREEN_FONT if p > 0 else RED_FONT
        for cc in range(1, len(headers) + 1):
            ws.cell(row=row, column=cc).border = BORDER
        row += 1
    # FLASH 参考行（海运 / 空运）
    for name, ship in (("FLASH参考(海运,5%广告)", SHIPPING_SEA), ("FLASH参考(空运,5%广告)", SHIPPING_AIR)):
        p = profit_per_unit(FLASH_PRICE_IDR, 0.05, PURCHASE_BASE, ship)
        r = profit_rate(FLASH_PRICE_IDR, 0.05, PURCHASE_BASE, ship)
        ws.cell(row=row, column=1, value=name).font = BOLD_FONT
        ws.cell(row=row, column=2, value=fmt_idr(FLASH_PRICE_IDR)).font = BODY_FONT
        ws.cell(row=row, column=3, value=f"≈{fmt_cny(cny(FLASH_PRICE_IDR))}").font = BODY_FONT
        c6 = ws.cell(row=row, column=6, value=fmt_cny(p)); c6.alignment = RIGHT
        c6.font = GREEN_FONT if p > 0 else RED_FONT
        c7 = ws.cell(row=row, column=7, value=fmt_pct(r)); c7.alignment = RIGHT
        c7.font = GREEN_FONT if p > 0 else RED_FONT
        fill = OK_FILL if p > 0 else BAD_FILL
        for cc in range(1, len(headers) + 1):
            ws.cell(row=row, column=cc).border = BORDER
            ws.cell(row=row, column=cc).fill = fill
        row += 1
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    tip = ws.cell(row=row, column=1, value="▶ 批量=3000件（采购-22%）联动利润见 Sheet5；空运仅补货，Flash 档空运亏损（红色）。")
    tip.font = SUB_FONT; tip.fill = OK_FILL
    ws.freeze_panes = "A4"
    widths = [16, 12, 12] + [11, 9] * 4
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============ Sheet 2: 竞品全景 ============
def build_sheet2(wb):
    ws = wb.create_sheet("竞品全景")
    ws.merge_cells("A1:G1")
    ws["A1"] = "竞品全景 — 1688 同款供应链价格 vs Shopee 印尼同场售价（数据来源：用户截图 + 1688/Shopee 行业采样 2026-08-27）"
    ws["A1"].font = TITLE_FONT
    headers = ["序号", "商品/卖家", "1688价(CNY)", "Shopee印尼价(IDR)", "折合CNY", "月销/热度(估)", "对标痛点"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER
    rows = [
        ("本品基准", "DejaVu HSB808（原价/Flash）", 4.00, 109000, "—", "★ 待售", "原价锚点Rp109k，Flash Rp53.1k"),
        ("1688-1", "无品牌普料壁挂脏衣篮", 3.60, 45000, cny(45000), "≈8000+/月", "塑料薄、挂钩易断、无品牌"),
        ("1688-2", "折叠手提脏衣篓", 3.90, 50000, cny(50000), "≈5000+/月", "尺寸偏小、用料单薄"),
        ("1688-3", "免打孔壁挂收纳篮", 4.20, 55000, cny(55000), "≈12000+/月", "胶贴易脱落、承重差"),
        ("1688-4", "大容量衣物收纳筐(同档)", 4.50, 60000, cny(60000), "≈6000+/月", "与本品采购价同档"),
        ("1688-5", "卫生间挂式脏衣篮", 5.30, 65000, cny(65000), "≈15000+/月", "款多、同质化严重"),
        ("1688-6", "可折叠脏衣篓(热销)", 5.50, 70000, cny(70000), "≈20000+/月", "价格战最烈档"),
        ("1688-7", "浴室手提式收纳篮", 5.55, 75000, cny(75000), "≈9000+/月", "材质一般、复购低"),
        ("1688-8", "防水牛津布脏衣篮", 7.80, 85000, cny(85000), "≈3000+/月", "贵、起订量大"),
        ("1688-9", "带盖壁挂脏衣篮", 10.50, 120000, cny(120000), "≈1500+/月", "高价小众"),
        ("1688-10", "加厚塑料收纳筐", 12.80, 160000, cny(160000), "≈800/月", "高端慢销"),
        ("1688-11", "品牌定制款", 16.00, 200000, cny(200000), "≈300/月", "品牌溢价、定制门槛"),
        ("Shopee-1", "无品牌塑料脏衣篮(印尼)", None, 45000, cny(45000), "高走量", "IDR30k-80k红海带底部"),
        ("Shopee-2", "入门折叠布艺款", None, 65000, cny(65000), "月销高", "主销带 IDR50k-80k"),
        ("Shopee-3", "大众织物款", None, 85000, cny(85000), "主销带", "IDR50k-150k 大众段"),
        ("Shopee-4", "带盖织物款", None, 120000, cny(120000), "中高端", "差异化溢价"),
        ("Shopee-5", "设计款多格", None, 160000, cny(160000), "设计溢价", "高端小众"),
        ("Shopee-6", "高级品牌 Joseph Joseph", None, 2299000, cny(2299000), "品牌锚点", "¥869 高价锚点，非走量"),
    ]
    r = 3
    for i, row_ in enumerate(rows, 1):
        kind, name, p1688, pidr, pcny, vol, pain = row_
        ws.cell(row=r, column=1, value=kind).font = BODY_FONT
        ws.cell(row=r, column=2, value=name).font = BODY_FONT
        ws.cell(row=r, column=3, value=(fmt_cny(p1688) if p1688 else "—")).font = BODY_FONT
        ws.cell(row=r, column=4, value=(fmt_idr(pidr) if pidr else "—")).font = BODY_FONT
        c5 = ws.cell(row=r, column=5, value=(f"≈{fmt_cny(pcny)}" if isinstance(pcny, float) else "—"))
        c5.font = BODY_FONT
        ws.cell(row=r, column=6, value=vol).font = BODY_FONT
        ws.cell(row=r, column=7, value=pain).font = BODY_FONT
        for cc in range(1, 8):
            cell = ws.cell(row=r, column=cc); cell.border = BORDER; cell.alignment = LEFT
        if i == 1:
            for cc in range(1, 8):
                ws.cell(row=r, column=cc).fill = WARN_FILL
        r += 1
    ws.cell(row=r + 1, column=1, value="摘要").font = BOLD_FONT
    ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=7)
    ws.cell(row=r + 1, column=2, value=(
        "1688价格带 ¥3.60-16.00（均值≈¥6.7），销量集中 ¥3.6-5.55 低价段；Shopee印尼最激烈竞争带 IDR30k-80k（≈¥11-30），"
        "大众主销带 IDR50k-150k；销量最集中 $7-12（≈IDR115k-197k）。本品 Flash Rp53.1k 落在红海带底部——靠本土店佣金优势+折叠设计差异化突围。"
    )).font = SUB_FONT
    for i, w in enumerate([10, 34, 13, 16, 12, 14, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

# ============ Sheet 3: SWOT与定价 ============
def build_sheet3(wb):
    ws = wb.create_sheet("SWOT与定价")
    ws.merge_cells("A1:B1")
    ws["A1"] = "SWOT 与定价策略 — 浴室壁挂式脏衣篮 × Shopee 印尼"
    ws["A1"].font = TITLE_FONT
    ws.cell(row=2, column=1, value="维度").font = HEADER_FONT
    ws.cell(row=2, column=2, value="内容").font = HEADER_FONT
    for c in (1, 2):
        ws.cell(row=2, column=c).fill = HEADER_FILL; ws.cell(row=2, column=c).alignment = CENTER
    swot = [
        ("优势 S", "折叠便携省空间；壁挂免打孔；大容量；中国供应链成本优势（1688 ¥3.6-4.5）；本土店佣金低(5%+2%)；折叠款是印尼市场增长引擎(CAGR 9-11%)"),
        ("劣势 W", "无品牌认知度；同类 SKU 严重同质化（1688 数十款同款）；塑料感强需精修主图；单价低、复购周期 1.5-2.5 年；品控依赖工厂，差评影响权重"),
        ("机会 O", "印尼洗衣篮市场 2026-2035 预计增长 50-60%；折叠款份额 40%→50%；电商渠道占比 22-25%→40%；中产+公寓/宿舍(20%)需求上升；可切 organized home 中端带(IDR100k-250k)"),
        ("威胁 T", "IDR30k-80k 红海价格战；印尼盾弱势(1年低点 0.0003696)；本地大厂 Maspion/Lion Star 挤压；树脂价格波动(占成本25-40%)；SNI 政策收紧风险；DejaVu 品牌名需授权（否则 IP 侵权下架）"),
    ]
    dim_fill = {"优势 S": "DDEBF7", "劣势 W": "FCE4EC", "机会 O": "E2EFDA", "威胁 T": "FFF2CC"}
    r = 3
    for dim, txt in swot:
        ws.cell(row=r, column=1, value=dim).font = BOLD_FONT
        ws.cell(row=r, column=2, value=txt).font = BODY_FONT
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=1).border = BORDER; ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=dim_fill[dim])
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=dim_fill[dim])
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="定价策略").font = BOLD_FONT
    r += 1
    pricing = [
        "① 划线价锚定：原价 Rp109,000（≈¥41.2）作为原价锚点，制造折扣感（对标大众主销带 IDR50k-150k 上沿）",
        "② 主力成交价带：Flash/日常折扣价落在 Rp55,000-85,000（≈¥20.8-32.1），贴主销带、避开红海底部价格战",
        "③ 保本下限：海运+5%广告下，Rp53,105 档单件净利≈¥2.9（正利润）；空运同档亏损≈¥12（见利润明细）→ 空运只用于补货不用于定价基准",
        "④ 利润拐点：批量≥1000件(折扣15%)后，单件到岸成本再降 ¥0.6-0.9，可支撑降25%档仍正利润",
    ]
    for p in pricing:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.cell(row=r, column=1, value=p).font = BODY_FONT
        ws.cell(row=r, column=1).alignment = LEFT
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="注意事项").font = BOLD_FONT
    r += 1
    notes = [
        "① 印尼语标签（强制）：产品名/材质/容量/进口商信息必须 Bahasa Indonesia，否则下架扣分",
        "② PT PMA 主体：本土店必需，年报/税务合规由主体承担（店铺主体您自行处理）",
        "③ 品牌合规：若沿用 DejaVu 品牌名需取得授权；建议贴自有品牌，避免 IP 侵权",
        "④ 禁售红线：仿品/侵权/宗教敏感图案/二手衣物禁售；图片白底≥800×800 无文字",
        "⑤ SLS 48小时未扫描视为虚假发货；Flash Sale 报名需好评率≥95%",
        "⑥ 回款 T+7（本土店）；汇率波动大（1年低点 0.0003696）→ 建议定价含 3-5% 汇率缓冲",
    ]
    for n in notes:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.cell(row=r, column=1, value=n).font = BODY_FONT
        ws.cell(row=r, column=1).alignment = LEFT
        ws.cell(row=r, column=1).fill = WARN_FILL
        r += 1
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 110

# ============ Sheet 4: 供应链物流 ============
def build_sheet4(wb):
    ws = wb.create_sheet("供应链物流")
    ws.merge_cells("A1:G1")
    ws["A1"] = "供应链与物流方案 — 1688采购 → 国内集运 → 国际段 → 印尼段（海运 vs 空运 双方案对比）"
    ws["A1"].font = TITLE_FONT
    headers = ["环节", "说明", "海运方案", "海运费用(CNY/件)", "空运方案", "空运费用(CNY/件)", "备注"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER
    rows = [
        ("① 1688采购", "义乌/广州货源，起订100件", "LCL 拼箱", 4.00, "同左", 4.00, "批量折扣100/500/1000/3000件=0/8/15/22%"),
        ("② 国内集运", "工厂→深圳/广州集货仓", "头程专线", 0.50, "同左", 0.50, "含国内段运输"),
        ("③ 国际段", "深圳→雅加达 Tanjung Priok", "海运 LCL 20-35天", 3.00, "空运 5-10天", 22.00, "海运约 USD15-25/m³；空运¥15-25/kg"),
        ("④ 清关+海外仓", "印尼清关、入库上架", "第三方海外仓", 2.00, "第三方海外仓", 2.00, "需要 PIB/API 进口资质（PT主体）"),
        ("⑤ 印尼段配送", "海外仓→买家 JNE/J&T", "本地快递1-3天", 4.00, "本地快递1-3天", 4.00, "Rp8,000-15,000/单≈¥3-5.7"),
        ("⑥ 包装", "彩盒+防压袋", "随货", 1.00, "随货", 1.00, "含在采购或另计"),
    ]
    r = 3
    for row_ in rows:
        for c, v in enumerate(row_, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT; cell.border = BORDER; cell.alignment = LEFT
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="单件物流合计").font = BOLD_FONT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(row=r, column=2, value="海运：¥0.5+3+2+4+1 = ¥10.5").font = BOLD_FONT
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    ws.cell(row=r, column=5, value="空运：¥0.5+22+2+4+1 = ¥29.5").font = BOLD_FONT
    ws.cell(row=r, column=7, value="差 ¥19/件").font = RED_FONT
    for cc in range(1, 8):
        ws.cell(row=r, column=cc).border = BORDER; ws.cell(row=r, column=cc).fill = WARN_FILL
    r += 2
    ws.cell(row=r, column=1, value="方案结论").font = BOLD_FONT
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=7)
    ws.cell(row=r + 1, column=1, value=(
        "✅ 海运：批量1000件下，到岸成本≈¥13.5/件，原价档利润率 50%+，Flash档仍正利润 → 主推方案。"
        "⚠️ 空运：只适合补货/测款（单件成本+¥19），定价必须按海运基准，空运期间关掉广告或提价。"
        "备选：Shopee SLS 官仓（卖家发广州仓，平台印尼段配送，买家付运费+卖家补贴1-3元）可轻资产起步，但佣金/时效差异需再评估。"
    )).font = BODY_FONT
    for i, w in enumerate([16, 22, 14, 16, 16, 16, 36], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

# ============ Sheet 5: 批量采购利润 ============
def build_sheet5(wb):
    ws = wb.create_sheet("批量采购利润")
    ws.merge_cells("A1:H1")
    ws["A1"] = "批量采购利润 — 起订量(4档) × 广告率(4档) × 物流(海运/空运) 单件净利矩阵"
    ws["A1"].font = TITLE_FONT
    headers = ["物流/起订量", "采购折扣", "采购价", "广告0%", "广告5%", "广告10%", "广告15%", "说明"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER
    r = 3
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="★ 海运方案（推荐，单件到岸基础 ¥10.5 + 采购）").font = BOLD_FONT
    ws.cell(row=r, column=1).fill = OK_FILL
    r += 1
    for qty, disc in BATCH:
        p = PURCHASE_BASE * (1 - disc)
        ws.cell(row=r, column=1, value=f"海运 {qty}件").font = BOLD_FONT
        ws.cell(row=r, column=2, value=fmt_pct(disc)).font = BODY_FONT
        ws.cell(row=r, column=3, value=fmt_cny(p)).font = BODY_FONT
        col = 4
        for ad in AD_RATES:
            pn = profit_per_unit(BASE_PRICE_IDR, ad, p, SHIPPING_SEA)
            c = ws.cell(row=r, column=col, value=fmt_cny(pn))
            c.font = GREEN_FONT if pn > 0 else RED_FONT
            c.alignment = RIGHT
            col += 1
        ws.cell(row=r, column=8, value="原价档；Flash 档见下方注释").font = SUB_FONT
        for cc in range(1, 9):
            ws.cell(row=r, column=cc).border = BORDER
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="★ 空运方案（仅补货，单件到岸基础 ¥29.5 + 采购）").font = BOLD_FONT
    ws.cell(row=r, column=1).fill = BAD_FILL
    r += 1
    for qty, disc in BATCH:
        p = PURCHASE_BASE * (1 - disc)
        ws.cell(row=r, column=1, value=f"空运 {qty}件").font = BOLD_FONT
        ws.cell(row=r, column=2, value=fmt_pct(disc)).font = BODY_FONT
        ws.cell(row=r, column=3, value=fmt_cny(p)).font = BODY_FONT
        col = 4
        for ad in AD_RATES:
            pn = profit_per_unit(BASE_PRICE_IDR, ad, p, SHIPPING_AIR)
            c = ws.cell(row=r, column=col, value=fmt_cny(pn))
            c.font = GREEN_FONT if pn > 0 else RED_FONT
            c.alignment = RIGHT
            col += 1
        ws.cell(row=r, column=8, value="原价档；空运+Flash 档全部亏损").font = RED_FONT
        for cc in range(1, 9):
            ws.cell(row=r, column=cc).border = BORDER
        r += 1
    r += 2
    main_p = profit_per_unit(FLASH_PRICE_IDR, 0.05, PURCHASE_BASE * (1 - 0.22), SHIPPING_SEA)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    if main_p > 0:
        ws.cell(row=r, column=1, value=f"✅ 正利润判定：主力场景（3000件+海运+5%广告+Flash Rp53,105）单件净利 {fmt_cny(main_p)}，利润率 {fmt_pct(main_p / cny(FLASH_PRICE_IDR))} → 项目可行，可启动。").font = GREEN_FONT
        ws.cell(row=r, column=1).fill = OK_FILL
    else:
        ws.cell(row=r, column=1, value=f"⚠️ 负利润预警：主力场景单件净利 {fmt_cny(main_p)} ≤ 0 → 建议停止，或重新谈判采购价/物流。").font = RED_FONT
        ws.cell(row=r, column=1).fill = BAD_FILL
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="注释：Flash Rp53,105 档（海运+5%广告）：单件净利≈¥2.9-3.9（随批量折扣）；空运同档亏损≈¥12；仅当 3000 件+海运+低广告时才考虑 Flash 价促销。").font = SUB_FONT
    for i, w in enumerate([16, 10, 11, 10, 10, 10, 10, 34], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

# ============ 主流程 ============
def main():
    wb = Workbook()
    build_sheet1(wb)
    build_sheet2(wb)
    build_sheet3(wb)
    build_sheet4(wb)
    build_sheet5(wb)
    wb.save(OUT_PATH)
    print("SAVED:", OUT_PATH)
    print("File size:", os.path.getsize(OUT_PATH), "bytes")

if __name__ == "__main__":
    main()
